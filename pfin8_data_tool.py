"""
P-Fin 8 Data Exploration Tool
Built with Streamlit

To run: streamlit run pfin8_data_tool.py
Required: streamlit, pandas, numpy, plotly, openpyxl
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from pathlib import Path

# ==============================================================================
# CONFIGURATION
# ==============================================================================
DEBUG_MODE = True
DATA_DIR = Path(".")  # Change to the folder containing the xlsx files

# ==============================================================================
# CONSTANTS
# ==============================================================================
TOPIC_NAMES = {
    "Earnings": "pfin8_earningsQCorrect",
    "Consuming": "pfin8_consumingQCorrect",
    "Savings": "pfin8_savingsQCorrect",
    "Investing": "pfin8_investingQCorrect",
    "Borrowing/Managing Debt": "pfin8_borrowingQCorrect",
    "Insuring": "pfin8_insuringQCorrect",
    "Comprehending Uncertainty": "pfin8_compreUncertQCorrect",
    "Go-to Information Sources": "pfin8_infoSourcesQCorrect",
}

TOPIC_CAT3_NAMES = {
    "Earnings": "pfin8_earnings_cat3",
    "Consuming": "pfin8_consuming_cat3",
    "Savings": "pfin8_savings_cat3",
    "Investing": "pfin8_investing_cat3",
    "Borrowing/Managing Debt": "pfin8_borrowing_cat3",
    "Insuring": "pfin8_insuring_cat3",
    "Comprehending Uncertainty": "pfin8_compreUncert_cat3",
    "Go-to Information Sources": "pfin8_infoSources_cat3",
}

DEMOGRAPHIC_VARIABLES = {
    "Age (Buckets)": "age_category",
    "Age (Custom Range)": "reported_age",
    "Generation": "generation_category",
    "Gender": "gender",
    "Race/Ethnicity": "race_ethnicity_category",
    "Marital Status": "marital_status",
    "Dependent Children Under 18": "has_dependent_children",
    "Education": "education_category",
    "Financial Education Course": "took_Financial_Education",
    "Employment Status": "employment_category",
    "Income": "income_category",
}

DEMOGRAPHIC_DISPLAY_LABELS = {
    "age_category": "Age",
    "reported_age": "Age",
    "generation_category": "Generation",
    "gender": "Gender",
    "race_ethnicity_category": "Race/Ethnicity",
    "marital_status": "Marital Status",
    "has_dependent_children": "Do you have dependent children under the age of 18?",
    "education_category": "Education",
    "took_Financial_Education": "Have you ever participated in a financial education class or program?",
    "employment_category": "Employment Status",
    "income_category": "Income",
}

FINANCIAL_WELLBEING_VARIABLES = {
    "Debt Constraint": "debt_constrained_responses",
    "Financial Fragility": "is_Financially_Fragile",
    "Non-Retirement Savings (One Month)": "suffretirement_savings_responses",
    "Time Thinking About Finances (Overall)": "time_thinking_finances",
    "Time Thinking About Finances (At Work)": "worktime_thinking_finances",
}

FINANCIAL_WELLBEING_LABELS = {
    "debt_constrained_responses": "Do debt and debt payments prevent you from adequately addressing other financial priorities?",
    "is_Financially_Fragile": "How confident are you that you could come up with $2,000 if an unexpected need arose within the next month?",
    "suffretirement_savings_responses": "Do you have non-retirement savings sufficient to cover one month of living expenses if needed?",
    "time_thinking_finances": "How much time (hours per week) do you typically spend thinking about and dealing with issues related to your personal finances?",
    "worktime_thinking_finances": "How many of these hours occur at work?",
}

TOTAL_CORRECT_LABELS = {
    0: "None Correct", 1: "1", 2: "2", 3: "3", 4: "4",
    5: "5", 6: "6", 7: "7", 8: "All Correct",
}

BINARY_DISPLAY = {0: "No", 1: "Yes"}

THINKING_TIME_ORDER = ["No hours", "1h", "2h", "3-4h", "5-9h", "10-19h", "20h+"]

AGE_BUCKET_ORDER = ["18-32", "33-47", "48-62", "63-79", "80+"]
GENERATION_ORDER = ["genZ", "genY", "genX", "boomer", "silent"]
EDUCATION_ORDER = ["Less than HS", "High School", "Some College", "Bachelor's degree or higher"]
INCOME_ORDER = ["<$25K", "$25-50K", "$50-100K", ">$100K"]

MIN_SAMPLE_SIZE = 30

# ==============================================================================
# DATA LOADING
# ==============================================================================
@st.cache_data
def load_all_years():
    df = pd.read_excel(DATA_DIR / "allYearsPFin8.xlsx")
    cat3_cols = [c for c in df.columns if "cat3" in c]
    for col in cat3_cols:
        df[col] = df[col].replace("Corrcet", "Correct")
    return df


@st.cache_data
def load_genpop():
    df = pd.read_excel(DATA_DIR / "PFin2026_GenPop.xlsx")
    cat3_cols = [c for c in df.columns if "cat3" in c]
    for col in cat3_cols:
        df[col] = df[col].replace("Corrcet", "Correct")
    # Fix "None" in thinking columns using openpyxl for raw read
    from openpyxl import load_workbook
    wb = load_workbook(DATA_DIR / "PFin2026_GenPop.xlsx")
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    for col_name in ["time_thinking_finances", "worktime_thinking_finances"]:
        col_idx = headers.index(col_name) + 1
        vals = []
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val == "None":
                vals.append("No hours")
            elif val == "" or val is None:
                vals.append(np.nan)
            else:
                vals.append(val)
        df[col_name] = vals
    # Map binary display for has_dependent_children and took_Financial_Education
    df["has_dependent_children_display"] = df["has_dependent_children"].map(BINARY_DISPLAY)
    df["took_Financial_Education_display"] = df["took_Financial_Education"].map(BINARY_DISPLAY)
    return df


# ==============================================================================
# SANITY CHECK FUNCTIONS
# ==============================================================================
def check_percentages_valid(percentages, label=""):
    issues = []
    for p in percentages:
        if pd.notna(p) and (p < 0 or p > 100):
            issues.append(f"Percentage out of range [0,100]: {p:.2f} in {label}")
    return issues


def check_group_counts(df, group_col, total_n):
    group_n = df.groupby(group_col).size().sum()
    if group_n != total_n:
        return [f"Group counts ({group_n}) don't sum to total ({total_n}) for {group_col}"]
    return []


def check_sample_size(n, label=""):
    if n < MIN_SAMPLE_SIZE:
        return f"⚠️ Small sample size (n={n}) for {label}. Results may be unreliable."
    return None


def run_sanity_checks(df, weight_col, result_data, environment, analysis_variable=None):
    checks = {"passed": [], "warnings": [], "errors": []}

    # Check weights are positive
    if (df[weight_col] <= 0).any():
        checks["errors"].append("Some survey weights are zero or negative")
    else:
        checks["passed"].append("All survey weights are positive")

    # Check total rows
    n = len(df)
    checks["passed"].append(f"Total observations in filtered data: {n}")

    # Check percentages in results
    if "percentage" in result_data.columns:
        pct_issues = check_percentages_valid(result_data["percentage"].values, "results")
        if pct_issues:
            checks["errors"].extend(pct_issues)
        else:
            checks["passed"].append("All percentages are within [0, 100]")

    # Check sample sizes per group
    if analysis_variable and analysis_variable in df.columns:
        for group, group_df in df.groupby(analysis_variable):
            warning = check_sample_size(len(group_df), str(group))
            if warning:
                checks["warnings"].append(warning)

    # Check that weighted percentages sum reasonably
    if "percentage" in result_data.columns:
        group_check_col = None
        cat_check_col = None
        for col in ["group_value", "group"]:
            if col in result_data.columns:
                group_check_col = col
                break
        for col in ["response_category", "category"]:
            if col in result_data.columns:
                cat_check_col = col
                break
        if group_check_col and cat_check_col:
            # Only check sum-to-100% when all response categories are present
            _all_resp = {"Correct", "Incorrect", "Don't Know"}
            _present_cats = set(result_data[cat_check_col].unique()) if cat_check_col in result_data.columns else set()
            _full_cats_present = not _present_cats or _present_cats >= _all_resp
            for grp_name, grp_data in result_data.groupby(group_check_col):
                total = grp_data.groupby(cat_check_col)["percentage"].sum().sum()
                # Only check if there's a facet (topic) or not
                if _full_cats_present and "topic" in result_data.columns:
                    for topic, topic_data in grp_data.groupby("topic"):
                        t = topic_data["percentage"].sum()
                        if abs(t - 100) > 1:
                            checks["warnings"].append(
                                f"Percentages for {grp_name}/{topic} sum to {t:.1f}%, expected ~100%"
                            )

    return checks


# ==============================================================================
# WEIGHTED CALCULATION FUNCTIONS
# ==============================================================================
def weighted_mean(df, col, weight_col="survey_weight"):
    valid = df.dropna(subset=[col, weight_col])
    if len(valid) == 0:
        return np.nan
    return np.average(valid[col], weights=valid[weight_col])


def weighted_percentage_binary(df, col, weight_col="survey_weight"):
    return weighted_mean(df, col, weight_col) * 100


def weighted_percentage_cat3(df, col, weight_col="survey_weight"):
    valid = df.dropna(subset=[col, weight_col])
    if len(valid) == 0:
        return pd.DataFrame()
    total_weight = valid[weight_col].sum()
    result = valid.groupby(col)[weight_col].sum() / total_weight * 100
    return result.reset_index().rename(columns={col: "category", weight_col: "percentage"})


def weighted_total_correct_distribution(df, weight_col="survey_weight"):
    valid = df.dropna(subset=["pfin8_totalCorrect", weight_col])
    if len(valid) == 0:
        return pd.DataFrame()
    total_weight = valid[weight_col].sum()
    result = valid.groupby("pfin8_totalCorrect")[weight_col].sum() / total_weight * 100
    result = result.reset_index().rename(
        columns={"pfin8_totalCorrect": "score", weight_col: "percentage"}
    )
    result["score_label"] = result["score"].map(TOTAL_CORRECT_LABELS)
    return result


# ==============================================================================
# CHART TYPE VALIDATION
# ==============================================================================
def get_valid_chart_types(analysis_type, view_mode, environment, axis_legend=None, n_legend_groups=1, n_total_correct=9, n_x_groups=2, n_response_cats=None):
    valid = []
    # Bar and Grouped Bar are mutually exclusive based on legend group count
    if n_legend_groups == 1:
        bar_option = "Bar Chart"
        h_bar_option = "Horizontal Bar Chart"
    else:
        bar_option = "Grouped Bar Chart"
        h_bar_option = "Horizontal Grouped Bar Chart"

    if analysis_type == "Topic Bucket":
        if view_mode == "3-Category (Correct / Incorrect / Don't Know)":
            valid = [bar_option, h_bar_option]
            # Stacked and pie only valid when all 3 response categories are selected
            all_cats_selected = n_response_cats is None or n_response_cats == 3
            if all_cats_selected and (axis_legend == "Response Category" or n_legend_groups == 1):
                valid.append("Stacked Bar Chart")
                valid.append("Horizontal Stacked Bar Chart")
                valid.append("Pie Chart")
        else:
            valid = [bar_option, h_bar_option]
            if n_x_groups > 1:
                valid.append("Line Chart")
    else:  # Number Correct
        valid = [bar_option, h_bar_option]
        if n_x_groups > 1:
            valid.append("Line Chart")
        # Stacked and pie only valid when full score range (0-8) is selected
        if n_total_correct >= 9:
            if axis_legend == "Number Correct" or n_legend_groups == 1:
                valid.append("Stacked Bar Chart")
                valid.append("Horizontal Stacked Bar Chart")
                valid.append("Pie Chart")
    # Table available unless it would be a single data point
    if not (n_x_groups == 1 and n_legend_groups == 1):
        valid.append("Table")
    return valid


# ==============================================================================
# CHART GENERATION
# ==============================================================================
def create_chart(chart_data, chart_type, title, x_label, y_label, color_col=None,
                 category_orders=None, group_label="group", hover_mode="binary",
                 legend_label=None, facet_col=None, n_legend_groups=None, pie_names_col=None,
                 show_pct_labels=False):
    fig = None
    try:
        label_map = {
            "x": x_label,
            "percentage": y_label,
            "group_value": group_label,
            "topic": "Topic",
            "response_category": "Response Category",
            "score_label": "Number Correct",
        }
        facet_args = {"facet_col": facet_col, "facet_col_wrap": 4} if facet_col else {}

        # Explicit color sequence matching Streamlit's default Plotly theme
        streamlit_colors = [
            "#636EFA", "#EF553B", "#00CC96", "#AB63FA", "#FFA15A",
            "#19D3F0", "#FF6692", "#B6E880", "#FF97FF", "#FECB52",
        ]

        if chart_type == "Bar Chart":
            fig = px.bar(
                chart_data, x="x", y="percentage",
                barmode="group",
                title=title, labels=label_map,
                category_orders=category_orders,
                color_discrete_sequence=streamlit_colors,
                **facet_args,
            )
            fig.update_layout(showlegend=False)
        elif chart_type == "Grouped Bar Chart":
            fig = px.bar(
                chart_data, x="x", y="percentage",
                color=color_col, barmode="group",
                title=title, labels=label_map,
                category_orders=category_orders,
                color_discrete_sequence=streamlit_colors,
                **facet_args,
            )
        elif chart_type == "Horizontal Bar Chart":
            fig = px.bar(
                chart_data, x="percentage", y="x",
                barmode="group",
                orientation="h",
                title=title, labels=label_map,
                category_orders=category_orders,
                color_discrete_sequence=streamlit_colors,
                **facet_args,
            )
            fig.update_layout(showlegend=False)
        elif chart_type == "Horizontal Grouped Bar Chart":
            fig = px.bar(
                chart_data, x="percentage", y="x",
                color=color_col, barmode="group",
                orientation="h",
                title=title, labels=label_map,
                category_orders=category_orders,
                color_discrete_sequence=streamlit_colors,
                **facet_args,
            )
        elif chart_type in ["Stacked Bar Chart", "Horizontal Stacked Bar Chart"]:
            fig = px.bar(
                chart_data,
                x="percentage" if chart_type == "Horizontal Stacked Bar Chart" else "x",
                y="x" if chart_type == "Horizontal Stacked Bar Chart" else "percentage",
                color=color_col, barmode="stack",
                orientation="h" if chart_type == "Horizontal Stacked Bar Chart" else "v",
                title=title, labels=label_map,
                category_orders=category_orders,
                color_discrete_sequence=streamlit_colors,
                **facet_args,
            )
        elif chart_type == "Pie Chart":
            # Determine which column has the parts-of-whole (slices)
            slice_col = pie_names_col if pie_names_col else color_col

            # Identify all non-slice dimensions that have multiple values
            facet_dims = []
            for col in ["x", "group_value", "topic", "score_label"]:
                if col in chart_data.columns and col != slice_col:
                    # Check if x is just a copy of this col
                    if col == "x":
                        continue
                    if chart_data[col].nunique() > 1:
                        facet_dims.append(col)
            # Also check x if it's not a copy of slice_col
            if "x" in chart_data.columns and not (
                slice_col in chart_data.columns and
                set(chart_data["x"].unique()) == set(chart_data[slice_col].unique())
            ):
                if chart_data["x"].nunique() > 1 and "x" not in facet_dims:
                    # Check x isn't a copy of an existing facet dim
                    is_copy = False
                    for fd in facet_dims:
                        if set(chart_data["x"].unique()) == set(chart_data[fd].unique()):
                            is_copy = True
                            break
                    if not is_copy:
                        facet_dims.insert(0, "x")

            if len(facet_dims) >= 2:
                # Use make_subplots to avoid px.pie's facet_row_spacing constraint
                chart_data["_pie_facet"] = chart_data[facet_dims[0]].astype(str) + " — " + chart_data[facet_dims[1]].astype(str)
                facet_labels = list(chart_data["_pie_facet"].unique())
                n_facets = len(facet_labels)
                n_cols = 4
                n_pie_rows = -(-n_facets // n_cols)
                # Consistent colors across all pies
                unique_slices = list(chart_data[slice_col].unique())
                color_map = {val: streamlit_colors[i % len(streamlit_colors)] for i, val in enumerate(unique_slices)}
                # Tight spacing that stays within Plotly's constraint
                v_spacing = min(0.02, 0.95 / (n_pie_rows - 1)) if n_pie_rows > 1 else 0
                fig = make_subplots(
                    rows=n_pie_rows, cols=n_cols,
                    specs=[[{"type": "pie"}] * n_cols for _ in range(n_pie_rows)],
                    subplot_titles=facet_labels + [""] * (n_pie_rows * n_cols - n_facets),
                    vertical_spacing=v_spacing,
                    horizontal_spacing=0.02,
                )
                for i, label in enumerate(facet_labels):
                    row_idx = i // n_cols + 1
                    col_idx = i % n_cols + 1
                    subset = chart_data[chart_data["_pie_facet"] == label]
                    sub_labels = subset[slice_col].tolist()
                    sub_colors = [color_map[lbl] for lbl in sub_labels]
                    fig.add_trace(
                        go.Pie(
                            values=subset["percentage"].tolist(),
                            labels=sub_labels,
                            name=label,
                            marker=dict(colors=sub_colors),
                            textposition="inside",
                            textinfo="percent+label",
                            textfont=dict(color="black"),
                            hovertemplate="%{label}: %{value:.0f}%<extra></extra>",
                            showlegend=(i == 0),
                        ),
                        row=row_idx, col=col_idx,
                    )
                fig.update_layout(title_text=title)
            elif len(facet_dims) == 1:
                fig = px.pie(
                    chart_data, values="percentage", names=slice_col,
                    title=title, labels=label_map,
                    color_discrete_sequence=streamlit_colors,
                    facet_col=facet_dims[0],
                    facet_col_wrap=4,
                )
            else:
                fig = px.pie(
                    chart_data, values="percentage", names=slice_col,
                    title=title, labels=label_map,
                    color_discrete_sequence=streamlit_colors,
                )
            if not n_legend_groups or n_legend_groups <= 10:
                fig.update_traces(textposition="inside", textinfo="percent+label", textfont=dict(color="black"))
        elif chart_type == "Line Chart":
            # If single group, don't color by legend
            if n_legend_groups and n_legend_groups <= 1:
                fig = px.line(
                    chart_data, x="x", y="percentage",
                    markers=True,
                    title=title, labels=label_map,
                    category_orders=category_orders,
                    color_discrete_sequence=streamlit_colors,
                    **facet_args,
                )
                fig.update_layout(showlegend=False)
            else:
                fig = px.line(
                    chart_data, x="x", y="percentage",
                    color=color_col, markers=True,
                    title=title, labels=label_map,
                    category_orders=category_orders,
                    color_discrete_sequence=streamlit_colors,
                    **facet_args,
                )

        if fig:
            # Determine if this is a single-group chart (no legend)
            no_legend_chart = chart_type in ["Bar Chart", "Horizontal Bar Chart"] or (
                chart_type == "Line Chart" and n_legend_groups and n_legend_groups <= 1
            )

            # Custom hover templates based on mode (skip for pie charts)
            if chart_type != "Pie Chart":
                if hover_mode == "cat3":
                    cat3_hover_labels = {
                        "Correct": "% Correct",
                        "Incorrect": "% Incorrect",
                        "Don't Know": "% Don't Know",
                    }
                    for trace in fig.data:
                        cat_name = trace.name
                        pct_label = cat3_hover_labels.get(cat_name, "% of Respondents")
                        if chart_type in ["Horizontal Bar Chart", "Horizontal Grouped Bar Chart"]:
                            trace.hovertemplate = (
                                f"{x_label}: %{{y}}<br>"
                                f"{pct_label}: %{{x:.0f}}%<br>"
                                f"<extra></extra>"
                            )
                        else:
                            trace.hovertemplate = (
                                f"{x_label}: %{{x}}<br>"
                                f"{pct_label}: %{{y:.0f}}%<br>"
                                f"<extra></extra>"
                            )
                elif hover_mode == "total_correct":
                    for trace in fig.data:
                        if no_legend_chart and chart_type == "Horizontal Bar Chart":
                            trace.hovertemplate = (
                                f"{x_label}: %{{y}}<br>"
                                f"% of Respondents: %{{x:.0f}}%<br>"
                                f"<extra></extra>"
                            )
                        elif no_legend_chart:
                            trace.hovertemplate = (
                                f"{x_label}: %{{x}}<br>"
                                f"% of Respondents: %{{y:.0f}}%<br>"
                                f"<extra></extra>"
                            )
                        elif chart_type == "Horizontal Grouped Bar Chart":
                            trace.hovertemplate = (
                                f"{x_label}: %{{y}}<br>"
                                f"% of Respondents: %{{x:.0f}}%<br>"
                                f"{group_label}: {trace.name}<br>"
                                f"<extra></extra>"
                            )
                        else:
                            trace.hovertemplate = (
                                f"{x_label}: %{{x}}<br>"
                                f"% of Respondents: %{{y:.0f}}%<br>"
                                f"{group_label}: {trace.name}<br>"
                                f"<extra></extra>"
                            )
                else:  # binary
                    for trace in fig.data:
                        if no_legend_chart and chart_type == "Horizontal Bar Chart":
                            trace.hovertemplate = (
                                f"{x_label}: %{{y}}<br>"
                                f"% Correct: %{{x:.0f}}%<br>"
                                f"<extra></extra>"
                            )
                        elif no_legend_chart:
                            trace.hovertemplate = (
                                f"{x_label}: %{{x}}<br>"
                                f"% Correct: %{{y:.0f}}%<br>"
                                f"<extra></extra>"
                            )
                        elif chart_type == "Horizontal Grouped Bar Chart":
                            trace.hovertemplate = (
                                f"{x_label}: %{{y}}<br>"
                                f"% Correct: %{{x:.0f}}%<br>"
                                f"{group_label}: {trace.name}<br>"
                                f"<extra></extra>"
                            )
                        else:
                            trace.hovertemplate = (
                                f"{x_label}: %{{x}}<br>"
                                f"% Correct: %{{y:.0f}}%<br>"
                                f"{group_label}: {trace.name}<br>"
                                f"<extra></extra>"
                            )

            # Layout adjustments
            if chart_type == "Pie Chart":
                n_pies = len(fig.data)
                pie_height = 400 if n_pies <= 4 else 400 * ((n_pies + 3) // 4)
                fig.update_layout(
                    legend_title_text=legend_label if legend_label else (group_label if color_col == "group" else (color_col if color_col else "")),
                    template="plotly_white",
                    font=dict(size=12, color="black"),
                    title_font=dict(size=16, color="black"),
                    height=pie_height,
                )
                # Clean up facet titles
                fig.for_each_annotation(lambda a: a.update(text=a.text.split("=")[-1]))
            elif chart_type in ["Horizontal Bar Chart", "Horizontal Grouped Bar Chart", "Horizontal Stacked Bar Chart"]:
                import math as _math
                _n_y_cats = chart_data["x"].nunique()
                if facet_col and facet_col in chart_data.columns:
                    _n_facets = chart_data[facet_col].nunique()
                    _n_grid_rows = _math.ceil(_n_facets / 4)
                    _h_bar_height = max(500, _n_grid_rows * max(200, _n_y_cats * 28 + 80) + 100)
                else:
                    _h_bar_height = max(400, _n_y_cats * 35 + 150)
                fig.update_layout(
                    yaxis_title=x_label,
                    xaxis_title=y_label,
                    legend_title_text=legend_label if legend_label else (group_label if color_col == "group" else (color_col if color_col else "")),
                    template="plotly_white",
                    font=dict(size=12, color="black"),
                    title_font=dict(size=16, color="black"),
                    height=_h_bar_height,
                )
                fig.update_xaxes(range=[0, 112 if show_pct_labels else 105])
                # Add top padding so facet labels don't cover the topmost bar
                fig.update_yaxes(range=[-0.5, _n_y_cats - 0.5 + 0.8])
                if facet_col:
                    fig.for_each_annotation(lambda a: a.update(text=a.text.split("=")[-1]))
            else:
                fig.update_layout(
                    yaxis_title=y_label,
                    xaxis_title=x_label,
                    legend_title_text=legend_label if legend_label else (group_label if color_col == "group" else (color_col if color_col else "")),
                    template="plotly_white",
                    font=dict(size=12, color="black"),
                    title_font=dict(size=16, color="black"),
                    height=800 if facet_col else 500,
                )
                fig.update_yaxes(range=[0, 112 if show_pct_labels else 105])

            # Force all axis text to black (plotly_white template overrides global font color)
            fig.update_xaxes(tickfont=dict(color="black"), title_font=dict(color="black"))
            fig.update_yaxes(tickfont=dict(color="black"), title_font=dict(color="black"))
            fig.update_layout(legend=dict(font=dict(color="black"), title_font=dict(color="black")))

            # Cap bar width when there are 4 or fewer x-axis categories
            # Only for stacked bars — use bargap to add space around bars
            if chart_type in ["Stacked Bar Chart", "Horizontal Stacked Bar Chart"]:
                n_x_categories = chart_data["x"].nunique()
                if n_x_categories == 1:
                    fig.update_layout(bargap=0.8)
                elif n_x_categories == 2:
                    fig.update_layout(bargap=0.7)
                elif n_x_categories == 3:
                    fig.update_layout(bargap=0.6)
                elif n_x_categories == 4:
                    fig.update_layout(bargap=0.5)

            # Clean up facet subplot titles (remove "topic=" prefix)
            if facet_col:
                fig.for_each_annotation(lambda a: a.update(text=a.text.split("=")[-1]))

            # Add percentage labels on bars
            if show_pct_labels and chart_type in ["Stacked Bar Chart", "Horizontal Stacked Bar Chart"]:
                # For stacked bars, threshold is segments per bar (n_legend_groups)
                if not n_legend_groups or n_legend_groups <= 10:
                    _tmpl = "%{x:.0f}%" if chart_type == "Horizontal Stacked Bar Chart" else "%{y:.0f}%"
                    fig.update_traces(
                        texttemplate=_tmpl,
                        textposition="inside",
                        insidetextfont=dict(color="black"),
                    )
            elif show_pct_labels and chart_type in ["Bar Chart", "Grouped Bar Chart",
                                                     "Horizontal Bar Chart", "Horizontal Grouped Bar Chart"]:
                total_bars = sum(len(trace.x) if hasattr(trace, 'x') and trace.x is not None else 0
                                 for trace in fig.data)
                if total_bars <= 70:
                    if chart_type in ["Horizontal Bar Chart", "Horizontal Grouped Bar Chart"]:
                        text_template = "%{x:.0f}%"
                    else:
                        text_template = "%{y:.0f}%"

                    fig.update_traces(
                        texttemplate=text_template,
                        textposition="inside",
                        insidetextfont=dict(color="black"),
                        textangle=-90,
                    )
        # Add source annotation at bottom right of every chart
        _cur_title = fig.layout.title.text or ""
        fig.update_layout(
            title=dict(
                text=f"{_cur_title}<br><sup><span style='color:gray;font-size:11px'>Source: TIAA G-Flec Personal Finance Index</span></sup>",
            )
        )
    except Exception as e:
        st.error(f"Could not create chart: {str(e)}")
        return None
    return fig


# ==============================================================================
# DESCRIPTIVE NOTE GENERATION
# ==============================================================================
def generate_note(environment, analysis_type, view_mode, selected_topics, selected_range,
                  analysis_variable, subgroups, n_obs, dataset_name, year_range=None):
    parts = []
    parts.append("**Data Source:** TIAA G-Flec Personal Finance Index")

    if environment == "Over the Years":
        yr_text = ", ".join(str(y) for y in sorted(year_range)) if year_range else "2017–2026"
        parts.append(f"**Years:** {yr_text}")
    else:
        parts.append("**Survey Year:** 2026")

    if analysis_type == "Topic Bucket":
        topics_text = ", ".join(selected_topics) if selected_topics else "All Topics"
        mode_text = "binary (correct vs. not correct)" if "Binary" in view_mode else "3-category (correct, incorrect, don't know)"
        parts.append(f"**Topics:** {topics_text}")
        parts.append(f"**View:** {mode_text}")
    else:
        range_text = ", ".join(TOTAL_CORRECT_LABELS[i] for i in sorted(selected_range)) if selected_range else "0–8"
        parts.append(f"**Number Correct Range:** {range_text}")

    if environment != "Over the Years" and analysis_variable:
        parts.append(f"**Analysis Variable:** {analysis_variable}")
        if subgroups:
            parts.append(f"**Subgroups:** {', '.join(str(s) for s in subgroups)}")

    parts.append(f"**Sample Size (filtered):** n={n_obs:,}")
    parts.append("All results are weighted to represent the population using survey weights. "
                  "Missing, refused, and inapplicable responses have been excluded.")
    return " | ".join(parts[:4]) + "\n\n" + " | ".join(parts[4:])


# ==============================================================================
# DATA PREPARATION FUNCTIONS
# ==============================================================================
def prepare_topic_binary_data(df, topics, group_col, group_label, weight_col="survey_weight"):
    rows = []
    for group_val, group_df in df.groupby(group_col):
        for topic_name, topic_col in topics.items():
            pct = weighted_percentage_binary(group_df, topic_col, weight_col)
            rows.append({
                "topic": topic_name,
                "group_value": str(group_val),
                "percentage": pct,
                "n": len(group_df.dropna(subset=[topic_col])),
            })
    return pd.DataFrame(rows)


def prepare_topic_cat3_data(df, topics, group_col, group_label, weight_col="survey_weight"):
    rows = []
    for group_val, group_df in df.groupby(group_col):
        for topic_name, cat3_col in topics.items():
            cat3_data = weighted_percentage_cat3(group_df, cat3_col, weight_col)
            for _, row in cat3_data.iterrows():
                rows.append({
                    "topic": topic_name,
                    "group_value": str(group_val),
                    "percentage": row["percentage"],
                    "response_category": row["category"],
                    "n": len(group_df.dropna(subset=[cat3_col])),
                })
    return pd.DataFrame(rows)


def prepare_total_correct_data(df, group_col, weight_col="survey_weight", score_range=None):
    rows = []
    for group_val, group_df in df.groupby(group_col):
        dist = weighted_total_correct_distribution(group_df, weight_col)
        if score_range:
            dist = dist[dist["score"].isin(score_range)]
        for _, row in dist.iterrows():
            rows.append({
                "score_label": row["score_label"],
                "score": row["score"],
                "percentage": row["percentage"],
                "group_value": str(group_val),
                "n": len(group_df.dropna(subset=["pfin8_totalCorrect"])),
            })
    return pd.DataFrame(rows)


def get_category_order(col_name):
    orders = {
        "age_category": AGE_BUCKET_ORDER,
        "generation_category": GENERATION_ORDER,
        "education_category": EDUCATION_ORDER,
        "income_category": INCOME_ORDER,
        "time_thinking_finances": THINKING_TIME_ORDER,
        "worktime_thinking_finances": THINKING_TIME_ORDER,
    }
    return orders.get(col_name)


# ==============================================================================
# SIDEBAR FILTERS
# ==============================================================================
def render_sidebar(df_years, df_genpop):
    with st.sidebar:
        st.title("P-Fin 8 Data Tool")

        # Remove box styling from expanders so they look like plain divider-separated sections
        st.markdown("""
<style>
section[data-testid="stSidebar"] [data-testid="stExpander"] details {
    border: none !important;
    box-shadow: none !important;
    background: transparent !important;
}
section[data-testid="stSidebar"] [data-testid="stExpander"]:first-of-type {
    margin-top: 2.25rem !important;
}
section[data-testid="stSidebar"] [data-testid="stExpander"] details summary {
    border-radius: 0 !important;
    padding: 0.4rem 0 !important;
    border-top: 1px solid rgba(49, 51, 63, 0.2) !important;
}
section[data-testid="stSidebar"] [data-testid="stExpanderDetails"] {
    padding: 0.5rem 0 0 0 !important;
    border: none !important;
}
/* Show scrollbar when hovering over the sidebar */
section[data-testid="stSidebar"]:hover *::-webkit-scrollbar {
    width: 8px !important;
    -webkit-appearance: none !important;
}
section[data-testid="stSidebar"]:hover *::-webkit-scrollbar-track {
    background: rgba(49, 51, 63, 0.1) !important;
    border-radius: 4px !important;
}
section[data-testid="stSidebar"]:hover *::-webkit-scrollbar-thumb {
    background: rgba(49, 51, 63, 0.4) !important;
    border-radius: 4px !important;
    min-height: 40px !important;
}
</style>
""", unsafe_allow_html=True)

        # Section 1: Analysis Type
        with st.expander("Analysis Type", expanded=True):
            analysis_type = st.radio(
                "Analysis Type",
                ["Topic Bucket", "Number Correct", "Distribution of Responses"],
                help="Analyze by individual topic questions or total correct score",
                label_visibility="collapsed",
            )

        # Section 2: View Mode / Number Correct Range
        view_mode = None
        selected_topics = None
        selected_range = None

        _sec2_title = "View Mode" if analysis_type in ("Topic Bucket", "Distribution of Responses") else "Number Correct Range"
        selected_response_cats = None
        dist_response_cat = None
        dist_range_mode = None
        dist_buckets = None
        dist_custom_ranges = None
        with st.expander(_sec2_title, expanded=True):
            if analysis_type == "Distribution of Responses":
                st.markdown("**Response Type**")
                dist_response_cat = st.radio(
                    "Response Type",
                    ["Correct", "Incorrect", "Don't Know"],
                    label_visibility="collapsed",
                )
                st.markdown("**Distribution Ranges**")
                dist_range_mode = st.radio(
                    "Distribution Ranges",
                    ["Buckets", "Custom Ranges"],
                    label_visibility="collapsed",
                )
                _dist_bucket_options = ["0-2 (<26%)", "3-4 (26%-50%)", "5-6 (51%-75%)", "7-8 (76%-100%)"]
                dist_buckets = None
                dist_custom_ranges = None
                if dist_range_mode == "Buckets":
                    dist_buckets = st.multiselect(
                        "Select Number of Questions Ranges",
                        _dist_bucket_options,
                        default=_dist_bucket_options,
                    )
                    if not dist_buckets:
                        st.warning("Please select at least one range.")
                else:
                    _dist_num_groups = st.selectbox(
                        "Number of groups",
                        [1, 2, 3, 4],
                        index=2,
                        key="dist_num_groups",
                    )
                    st.markdown("**Define your groups**")
                    st.caption("Min: 0 · Max: 8")
                    _dist_groups = []
                    _dist_errors = []
                    for _i in range(_dist_num_groups):
                        st.markdown(f"**Number of Questions Group {_i+1}:**")
                        _dc1, _dc2 = st.columns(2)
                        with _dc1:
                            _ds = st.number_input(
                                "Start", min_value=0, max_value=8,
                                value=min(_i * (8 // _dist_num_groups), 8),
                                key=f"dist_start_{_i}",
                            )
                        with _dc2:
                            _de_default = min((_i + 1) * (8 // _dist_num_groups) - 1, 8) if _i < _dist_num_groups - 1 else 8
                            _de = st.number_input(
                                "End", min_value=0, max_value=8,
                                value=_de_default,
                                key=f"dist_end_{_i}",
                            )
                        if _de < _ds:
                            st.markdown(f'<p style="color: red; font-size: 0.85rem; margin: -10px 0 5px 0;">⚠️ Number of Questions Group {_i+1}: end must be ≥ start</p>', unsafe_allow_html=True)
                            _dist_errors.append(f"Number of Questions Group {_i+1}: end < start")
                        _dist_groups.append((_ds, _de))
                    for _i in range(len(_dist_groups)):
                        for _j in range(_i + 1, len(_dist_groups)):
                            _g1s, _g1e = _dist_groups[_i]
                            _g2s, _g2e = _dist_groups[_j]
                            if _g1s <= _g2e and _g2s <= _g1e:
                                _ov_s, _ov_e = max(_g1s, _g2s), min(_g1e, _g2e)
                                st.markdown(f'<p style="color: red; font-size: 0.85rem; margin: 0 0 5px 0;">⚠️ Number of Questions Groups {_i+1} and {_j+1} overlap ({_ov_s}–{_ov_e})</p>', unsafe_allow_html=True)
                                _dist_errors.append(f"Number of Questions Groups {_i+1} and {_j+1} overlap")
                    if _dist_errors:
                        st.error("Invalid groups — please adjust ranges")
                    _dist_labels = []
                    for _i, (_s, _e) in enumerate(_dist_groups):
                        if _i == len(_dist_groups) - 1 and _e == 8:
                            _dist_labels.append(f"{_s}+")
                        else:
                            _dist_labels.append(f"{_s}-{_e}")
                    dist_custom_ranges = {
                        "groups": _dist_groups,
                        "labels": _dist_labels,
                        "errors": _dist_errors,
                    }
            elif analysis_type == "Topic Bucket":
                view_mode = st.radio(
                    "View Mode",
                    ["Binary (Correct / Not Correct)", "3-Category (Correct / Incorrect / Don't Know)"],
                    label_visibility="collapsed",
                )
                if "Binary" in view_mode:
                    binary_response = st.radio(
                        "Response",
                        ["Correct", "Not Correct"],
                        help="Not Correct includes both Incorrect and Don't Know responses",
                    )
                    selected_response_cats = [binary_response]
                else:
                    _all_resp_cats = ["Correct", "Incorrect", "Don't Know"]
                    st.markdown("**Response Categories**")
                    selected_response_cats = st.multiselect(
                        "Response Categories",
                        _all_resp_cats,
                        default=_all_resp_cats,
                        label_visibility="collapsed",
                    )
                    if not selected_response_cats:
                        st.warning("Please select at least one response category.")
                all_topics = list(TOPIC_NAMES.keys())
                selected_topics = st.multiselect(
                    "Select Topics",
                    all_topics,
                    default=all_topics,
                    help="Choose which P-Fin 8 topics to include",
                )
                if not selected_topics:
                    st.warning("Please select at least one topic.")
            elif analysis_type == "Number Correct":
                _all_scores = list(range(9))
                for _s in _all_scores:
                    if f"nc_cb_{_s}" not in st.session_state:
                        st.session_state[f"nc_cb_{_s}"] = True
                _nc_btn1, _nc_btn2 = st.columns(2)
                with _nc_btn1:
                    if st.button("Select All", key="nc_select_all", use_container_width=True):
                        for _s in _all_scores:
                            st.session_state[f"nc_cb_{_s}"] = True
                with _nc_btn2:
                    if st.button("Deselect All", key="nc_deselect_all", use_container_width=True):
                        for _s in _all_scores:
                            st.session_state[f"nc_cb_{_s}"] = False
                _nc_col1, _nc_col2 = st.columns(2)
                selected_range = []
                for _i, _s in enumerate(_all_scores):
                    with _nc_col1 if _i % 2 == 0 else _nc_col2:
                        if st.checkbox(TOTAL_CORRECT_LABELS[_s], key=f"nc_cb_{_s}"):
                            selected_range.append(_s)
                if not selected_range:
                    st.warning("Please select at least one score.")

        # Section 3: Exploration Type
        with st.expander("Exploration Type", expanded=True):
            environment = st.radio(
                "Exploration Type",
                ["Over the Years", "Demographics", "Financial Well-Being"],
                help="Choose how you want to explore the P-Fin 8 data",
                label_visibility="collapsed",
            )

        # Section 4: Environment-specific filters
        analysis_variable = None
        analysis_col = None
        subgroups = None
        selected_years = None
        custom_age_range = None

        if environment == "Over the Years":
            _sec4_title = "Years"
        elif environment == "Demographics":
            _sec4_title = "Demographic Variable"
        else:
            _sec4_title = "Financial Well-Being Variable"

        with st.expander(_sec4_title, expanded=True):
            if environment == "Over the Years":
                _all_years = sorted([int(y) for y in df_years["survey_year"].unique()])
                # Initialize checkboxes to True on first load
                for _y in _all_years:
                    if f"year_cb_{_y}" not in st.session_state:
                        st.session_state[f"year_cb_{_y}"] = True
                # Select All / Deselect All buttons
                _btn_col1, _btn_col2 = st.columns(2)
                with _btn_col1:
                    if st.button("Select All", key="year_select_all", use_container_width=True):
                        for _y in _all_years:
                            st.session_state[f"year_cb_{_y}"] = True
                with _btn_col2:
                    if st.button("Deselect All", key="year_deselect_all", use_container_width=True):
                        for _y in _all_years:
                            st.session_state[f"year_cb_{_y}"] = False
                # 2-column checkbox grid
                _cb_col1, _cb_col2 = st.columns(2)
                selected_years = []
                for _i, _y in enumerate(_all_years):
                    with _cb_col1 if _i % 2 == 0 else _cb_col2:
                        if st.checkbox(str(_y), key=f"year_cb_{_y}"):
                            selected_years.append(_y)
                if not selected_years:
                    st.warning("Please select at least one year.")

            elif environment == "Demographics":
                analysis_variable = st.selectbox(
                    "Demographic Variable",
                    list(DEMOGRAPHIC_VARIABLES.keys()),
                    label_visibility="collapsed",
                )
                analysis_col = DEMOGRAPHIC_VARIABLES[analysis_variable]

                if analysis_variable == "Age (Custom Range)":
                    min_age = int(df_genpop["reported_age"].min())
                    max_age = int(df_genpop["reported_age"].max())

                    num_groups = st.selectbox(
                        "Number of age groups",
                        [1, 2, 3, 4, 5, 6, 7, 8],
                        index=2,
                    )

                    st.markdown("**Define your age groups**")
                    st.caption(f"Min age: {min_age} · Max age: {max_age}")
                    custom_age_groups = []
                    age_errors = []

                    for i in range(num_groups):
                        st.markdown(f"**Age Group {i+1}:**")
                        col1, col2 = st.columns(2)
                        with col1:
                            start = st.number_input(
                                "Start",
                                min_value=min_age, max_value=max_age,
                                value=min(min_age + i * ((max_age - min_age) // num_groups), max_age),
                                key=f"age_start_{i}",
                            )
                        with col2:
                            default_end = min(min_age + (i + 1) * ((max_age - min_age) // num_groups) - 1, max_age)
                            if i == num_groups - 1:
                                default_end = max_age
                            end = st.number_input(
                                "End",
                                min_value=min_age, max_value=max_age,
                                value=default_end,
                                key=f"age_end_{i}",
                            )

                        # Validate: end must be >= start
                        if end < start:
                            st.markdown(f'<p style="color: red; font-size: 0.85rem; margin: -10px 0 5px 0;">⚠️ Age Group {i+1}: end age must be ≥ start age</p>', unsafe_allow_html=True)
                            age_errors.append(f"Age Group {i+1}: end < start")

                        custom_age_groups.append((start, end))

                    # Validate: check for overlaps
                    for i in range(len(custom_age_groups)):
                        for j in range(i + 1, len(custom_age_groups)):
                            g1_start, g1_end = custom_age_groups[i]
                            g2_start, g2_end = custom_age_groups[j]
                            if g1_start <= g2_end and g2_start <= g1_end:
                                overlap_start = max(g1_start, g2_start)
                                overlap_end = min(g1_end, g2_end)
                                st.markdown(f'<p style="color: red; font-size: 0.85rem; margin: 0 0 5px 0;">⚠️ Age Groups {i+1} and {j+1} overlap (ages {overlap_start}–{overlap_end})</p>', unsafe_allow_html=True)
                                age_errors.append(f"Age Groups {i+1} and {j+1} overlap")

                    if age_errors:
                        st.error("Invalid groups — please adjust ranges")

                    # Build labels
                    custom_age_labels = []
                    for i, (s, e) in enumerate(custom_age_groups):
                        if i == len(custom_age_groups) - 1 and e == max_age:
                            custom_age_labels.append(f"{s}+")
                        else:
                            custom_age_labels.append(f"{s}-{e}")

                    custom_age_range = {
                        "groups": custom_age_groups,
                        "labels": custom_age_labels,
                        "errors": age_errors,
                    }
                elif analysis_variable == "Dependent Children Under 18":
                    subgroups = st.multiselect(
                        "Select Groups",
                        ["Yes", "No"],
                        default=["Yes", "No"],
                    )
                elif analysis_variable == "Financial Education Course":
                    subgroups = st.multiselect(
                        "Select Groups",
                        ["Yes", "No"],
                        default=["Yes", "No"],
                    )
                else:
                    available_values = sorted(df_genpop[analysis_col].dropna().unique().tolist())
                    order = get_category_order(analysis_col)
                    if order:
                        available_values = [v for v in order if v in available_values]
                    subgroups = st.multiselect(
                        f"Select {analysis_variable} Groups",
                        available_values,
                        default=available_values,
                    )

            elif environment == "Financial Well-Being":
                analysis_variable = st.selectbox(
                    "Financial Well-Being Variable",
                    list(FINANCIAL_WELLBEING_VARIABLES.keys()),
                    label_visibility="collapsed",
                )
                analysis_col = FINANCIAL_WELLBEING_VARIABLES[analysis_variable]
                available_values = df_genpop[analysis_col].dropna().unique().tolist()
                order = get_category_order(analysis_col)
                if order:
                    available_values = [v for v in order if v in available_values]
                else:
                    available_values = sorted(available_values, key=str)
                subgroups = st.multiselect(
                    f"Select Groups",
                    available_values,
                    default=available_values,
                )

        # Axis assignment
        # Determine the group dimension label
        if environment == "Over the Years":
            group_dim_label = "Year"
        elif environment == "Demographics":
            group_dim_label = analysis_variable if analysis_variable else "Demographic"
        else:
            group_dim_label = analysis_variable if analysis_variable else "Financial Well-Being"

        # Compute dimension sizes to determine if axis assignment should be shown
        n_topics = len(selected_topics) if selected_topics else 8
        n_total_correct = len(selected_range) if selected_range else 9

        # Count group dimension values
        if environment == "Over the Years":
            n_group = len(selected_years) if selected_years else 10
        elif analysis_variable == "Age (Custom Range)" and custom_age_range:
            n_group = len(custom_age_range.get("groups", []))
        elif subgroups:
            n_group = len(subgroups)
        else:
            n_group = 1

        axis_x = None
        axis_legend = None
        axis_facet = None
        single_group_value = None
        axis_assignment_shown = False
        _aa_info = None  # Info for deferred Axis Assignment expander
        n_response_cats = len(selected_response_cats) if selected_response_cats else (3 if view_mode and "3-Category" in view_mode else 1)

        if analysis_type == "Topic Bucket" and view_mode and "3-Category" in view_mode:
            if n_response_cats == 1:
                # Single response category selected — treat as 2-way (Topic vs group)
                axis_facet = None
                if n_topics == 1 and n_group == 1:
                    axis_x = "Topic"
                    axis_legend = group_dim_label
                    single_group_value = selected_topics[0] if selected_topics else None
                elif n_topics == 1 and n_group > 1:
                    axis_x = group_dim_label
                    axis_legend = "Topic"
                    single_group_value = selected_topics[0] if selected_topics else None
                elif n_group == 1 and n_topics > 1:
                    axis_x = "Topic"
                    axis_legend = group_dim_label
                    if environment == "Over the Years" and selected_years and len(selected_years) == 1:
                        single_group_value = str(selected_years[0])
                    elif subgroups and len(subgroups) == 1:
                        single_group_value = str(subgroups[0])
                else:
                    axis_assignment_shown = True
                    _dim2 = ["Topic", group_dim_label]
                    _aa_info = {'type': 'two_way', 'options': _dim2, 'default_idx': 0}
                    _curr_x = st.session_state.get('pfin8_aa_x', _dim2[0])
                    axis_x = _curr_x if _curr_x in _dim2 else _dim2[0]
                    axis_legend = [d for d in _dim2 if d != axis_x][0]
            else:
                dimensions = ["Topic", group_dim_label, "Response Category"]

                if n_topics == 1 and n_group == 1:
                    axis_x = "Response Category"
                    axis_legend = "Topic"
                    axis_facet = group_dim_label
                    single_group_value = selected_topics[0] if selected_topics else None
                elif n_topics == 1 and n_group > 1:
                    single_group_value = selected_topics[0] if selected_topics else None
                    remaining = [d for d in dimensions if d != "Topic"]
                    axis_assignment_shown = True
                    _aa_info = {'type': 'two_way', 'options': remaining, 'default_idx': 0}
                    _curr_x = st.session_state.get('pfin8_aa_x', remaining[0])
                    axis_x = _curr_x if _curr_x in remaining else remaining[0]
                    axis_legend = [d for d in remaining if d != axis_x][0]
                    axis_facet = "Topic"
                elif n_group == 1 and n_topics > 1:
                    if environment == "Over the Years" and selected_years and len(selected_years) == 1:
                        single_group_value = str(selected_years[0])
                    elif subgroups and len(subgroups) == 1:
                        single_group_value = str(subgroups[0])
                    remaining = [d for d in dimensions if d != group_dim_label]
                    axis_assignment_shown = True
                    _aa_info = {'type': 'two_way', 'options': remaining, 'default_idx': 0}
                    _curr_x = st.session_state.get('pfin8_aa_x', remaining[0])
                    axis_x = _curr_x if _curr_x in remaining else remaining[0]
                    axis_legend = [d for d in remaining if d != axis_x][0]
                    axis_facet = group_dim_label
                else:
                    axis_assignment_shown = True
                    _aa_info = {'type': 'three_way', 'options': dimensions, 'default_x_idx': 1}
                    _curr_x = st.session_state.get('pfin8_aa_x', dimensions[1])
                    axis_x = _curr_x if _curr_x in dimensions else dimensions[1]
                    _rem = [d for d in dimensions if d != axis_x]
                    _curr_legend = st.session_state.get('pfin8_aa_legend', _rem[0])
                    axis_legend = _curr_legend if _curr_legend in _rem else _rem[0]
                    axis_facet = [d for d in dimensions if d != axis_x and d != axis_legend][0]

        elif analysis_type == "Topic Bucket":
            if n_topics == 1 and n_group > 1:
                axis_x = group_dim_label
                axis_legend = "Topic"
                single_group_value = selected_topics[0] if selected_topics else None
            elif n_group == 1 and n_topics > 1:
                axis_x = "Topic"
                axis_legend = group_dim_label
                if environment == "Over the Years" and selected_years and len(selected_years) == 1:
                    single_group_value = str(selected_years[0])
                elif subgroups and len(subgroups) == 1:
                    single_group_value = str(subgroups[0])
            elif n_topics == 1 and n_group == 1:
                axis_x = "Topic"
                axis_legend = group_dim_label
                single_group_value = selected_topics[0] if selected_topics else None
            else:
                axis_assignment_shown = True
                _dim2 = ["Topic", group_dim_label]
                _aa_info = {'type': 'two_way', 'options': _dim2, 'default_idx': 0}
                _curr_x = st.session_state.get('pfin8_aa_x', _dim2[0])
                axis_x = _curr_x if _curr_x in _dim2 else _dim2[0]
                axis_legend = [d for d in _dim2 if d != axis_x][0]

        else:
            if n_total_correct == 1 and n_group > 1:
                axis_x = group_dim_label
                axis_legend = "Number Correct"
                single_group_value = TOTAL_CORRECT_LABELS.get(selected_range[0], str(selected_range[0])) if selected_range and len(selected_range) == 1 else None
            elif n_group == 1 and n_total_correct > 1:
                axis_x = "Number Correct"
                axis_legend = group_dim_label
                if environment == "Over the Years" and selected_years and len(selected_years) == 1:
                    single_group_value = str(selected_years[0])
                elif subgroups and len(subgroups) == 1:
                    single_group_value = str(subgroups[0])
            elif n_total_correct == 1 and n_group == 1:
                axis_x = "Number Correct"
                axis_legend = group_dim_label
            else:
                axis_assignment_shown = True
                _dim2 = ["Number Correct", group_dim_label]
                _aa_info = {'type': 'two_way', 'options': _dim2, 'default_idx': 0}
                _curr_x = st.session_state.get('pfin8_aa_x', _dim2[0])
                axis_x = _curr_x if _curr_x in _dim2 else _dim2[0]
                axis_legend = [d for d in _dim2 if d != axis_x][0]

        # Compute n_legend_groups and n_x_groups for chart type validation
        n_legend_groups = 1
        if axis_legend == "Topic":
            n_legend_groups = n_topics
        elif axis_legend == "Response Category":
            n_legend_groups = n_response_cats
        elif axis_legend == "Number Correct":
            n_legend_groups = n_total_correct
        elif axis_legend == group_dim_label:
            n_legend_groups = n_group

        n_x_groups = 1
        if axis_x == "Topic":
            n_x_groups = n_topics
        elif axis_x == "Response Category":
            n_x_groups = n_response_cats
        elif axis_x == "Number Correct":
            n_x_groups = n_total_correct
        elif axis_x == group_dim_label:
            n_x_groups = n_group

        # Section 5: Axis Assignment
        if axis_assignment_shown and _aa_info:
            with st.expander("Axis Assignment", expanded=True):
                _opts = _aa_info['options']
                if _aa_info['type'] == 'three_way':
                    axis_x = st.selectbox(
                        "X-Axis", _opts,
                        index=_opts.index(axis_x) if axis_x in _opts else _aa_info['default_x_idx'],
                        key='pfin8_aa_x',
                    )
                    _rem = [d for d in _opts if d != axis_x]
                    axis_legend = st.selectbox(
                        "Legend", _rem,
                        index=_rem.index(axis_legend) if axis_legend in _rem else 0,
                        key='pfin8_aa_legend',
                    )
                    axis_facet = [d for d in _opts if d != axis_x and d != axis_legend][0]
                    st.caption(f"Facet (panels): **{axis_facet}**")
                else:
                    axis_x = st.selectbox(
                        "X-Axis", _opts,
                        index=_opts.index(axis_x) if axis_x in _opts else _aa_info['default_idx'],
                        key='pfin8_aa_x',
                    )
                    axis_legend = [d for d in _opts if d != axis_x][0]
                    st.caption(f"Legend: **{axis_legend}**")

        # Section 6: Chart Type + Show percentages toggle
        with st.expander("Chart Type", expanded=True):
            valid_charts = get_valid_chart_types(analysis_type, view_mode, environment, axis_legend, n_legend_groups, n_total_correct, n_x_groups, n_response_cats)
            chart_type = st.selectbox("Chart Type", valid_charts, label_visibility="collapsed")

            show_pct_labels = False
            if chart_type in ["Bar Chart", "Grouped Bar Chart", "Horizontal Bar Chart",
                              "Horizontal Grouped Bar Chart", "Stacked Bar Chart", "Horizontal Stacked Bar Chart"]:
                show_pct_labels = st.toggle("Show percentages on bars", value=True)

        return {
            "environment": environment,
            "analysis_type": analysis_type,
            "view_mode": view_mode,
            "selected_topics": selected_topics,
            "selected_range": selected_range,
            "analysis_variable": analysis_variable,
            "analysis_col": analysis_col,
            "subgroups": subgroups,
            "selected_years": selected_years,
            "custom_age_range": custom_age_range,
            "chart_type": chart_type,
            "axis_x": axis_x,
            "axis_legend": axis_legend,
            "axis_facet": axis_facet,
            "group_dim_label": group_dim_label,
            "single_group_value": single_group_value,
            "show_pct_labels": show_pct_labels,
            "selected_response_cats": selected_response_cats,
            "dist_response_cat": dist_response_cat,
            "dist_range_mode": dist_range_mode,
            "dist_buckets": dist_buckets,
            "dist_custom_ranges": dist_custom_ranges,
        }


# ==============================================================================
# MAIN ANALYSIS AND VISUALIZATION
# ==============================================================================
def run_analysis(config, df_years, df_genpop):
    environment = config["environment"]
    analysis_type = config["analysis_type"]
    view_mode = config["view_mode"]
    chart_type = config["chart_type"]

    # Select dataset and apply filters
    if environment == "Over the Years":
        df = df_years.copy()
        selected_years = config["selected_years"]
        if not selected_years:
            return None, None, None, None, None
        df = df[df["survey_year"].isin(selected_years)]
        group_col = "survey_year"
        group_label = "Year"
        dataset_name = "allYearsPFin8 (2017–2026)"
        analysis_col = "survey_year"
    else:
        df = df_genpop.copy()
        analysis_col = config["analysis_col"]
        dataset_name = "PFin2026_GenPop (2026)"

        if config["analysis_variable"] == "Age (Custom Range)":
            age_config = config["custom_age_range"]
            if age_config and age_config.get("errors"):
                st.error("Invalid groups — please adjust ranges")
                return None, None, None, None, None

            groups = age_config["groups"]
            labels = age_config["labels"]

            # Assign each respondent to a group based on their age
            def assign_age_group(age):
                for i, (start, end) in enumerate(groups):
                    if start <= age <= end:
                        return labels[i]
                return None  # Age not in any group

            df["custom_age_group"] = df["reported_age"].apply(assign_age_group)
            df = df.dropna(subset=["custom_age_group"])

            group_col = "custom_age_group"
            group_label = "Age Range"
            analysis_col = "custom_age_group"

            # Set category order to match the label order (used later via config)
        elif config["analysis_variable"] == "Dependent Children Under 18":
            df["has_dependent_children_display"] = df["has_dependent_children"].map(BINARY_DISPLAY)
            group_col = "has_dependent_children_display"
            group_label = DEMOGRAPHIC_DISPLAY_LABELS.get("has_dependent_children", config["analysis_variable"])
            analysis_col = "has_dependent_children_display"
            if config["subgroups"]:
                df = df[df[group_col].isin(config["subgroups"])]
        elif config["analysis_variable"] == "Financial Education Course":
            df["took_Financial_Education_display"] = df["took_Financial_Education"].map(BINARY_DISPLAY)
            group_col = "took_Financial_Education_display"
            group_label = DEMOGRAPHIC_DISPLAY_LABELS.get("took_Financial_Education", config["analysis_variable"])
            analysis_col = "took_Financial_Education_display"
            if config["subgroups"]:
                df = df[df[group_col].isin(config["subgroups"])]
        else:
            group_col = analysis_col
            if environment == "Demographics":
                group_label = DEMOGRAPHIC_DISPLAY_LABELS.get(analysis_col, config["analysis_variable"])
            else:
                group_label = FINANCIAL_WELLBEING_LABELS.get(analysis_col, config["analysis_variable"])

            # Filter by subgroups
            if config["subgroups"]:
                df = df[df[group_col].isin(config["subgroups"])]

        # Drop missing values in the group column
        df = df.dropna(subset=[group_col])

    # Check if data remains after filtering
    if len(df) == 0:
        st.warning("No data available for the selected filters. Please adjust your selections.")
        return None, None, None, None, None

    # Determine category orders for the group column
    cat_order = get_category_order(group_col)
    category_orders = {}
    if group_col == "custom_age_group" and config.get("custom_age_range"):
        # Use the user-defined label order for custom age groups
        category_orders["group_value"] = config["custom_age_range"]["labels"]
    elif cat_order:
        available = [v for v in cat_order if v in df[group_col].unique()]
        category_orders["group_value"] = [str(v) for v in available]
    if environment == "Over the Years":
        _yr_order = sorted(selected_years) if selected_years else sorted(df[group_col].unique())
        category_orders["group_value"] = [str(v) for v in _yr_order]

    # Build chart data
    chart_data = None
    color_col = "group_value"
    x_label = ""
    hover_mode = "binary"
    use_facet = None
    group_dim_label = config["group_dim_label"]
    axis_x = config["axis_x"]
    axis_legend = config["axis_legend"]
    axis_facet = config.get("axis_facet")
    single_group_value = config.get("single_group_value")

    # Map dimension names to data columns
    def dim_to_col(dim_name, mode="binary"):
        if dim_name == "Topic":
            return "topic"
        elif dim_name == "Number Correct":
            return "score_label"
        elif dim_name == "Response Category":
            return "response_category"
        else:  # group dimension
            return "group_value"

    if analysis_type == "Topic Bucket":
        selected_topics = config["selected_topics"]
        if not selected_topics:
            return None, None, None, None, None

        selected_response_cats = config.get("selected_response_cats")
        if view_mode and "3-Category" in view_mode and selected_response_cats is not None and not selected_response_cats:
            return None, None, None, None, None

        if view_mode and "Binary" in view_mode:
            topics_map = {k: v for k, v in TOPIC_NAMES.items() if k in selected_topics}
            chart_data = prepare_topic_binary_data(df, topics_map, group_col, group_label)
            hover_mode = "binary"
            # Flip percentage if "Not Correct" is selected
            _show_not_correct = selected_response_cats and selected_response_cats == ["Not Correct"]
            if _show_not_correct:
                chart_data["percentage"] = 100 - chart_data["percentage"]
                y_label = "% Not Correct"
                _metric_label = "% Not Correct"
            else:
                y_label = "% Correct"
                _metric_label = "% Correct"

            # Assign axes
            x_col = dim_to_col(axis_x)
            legend_col = dim_to_col(axis_legend)
            x_dim_label = axis_x if axis_x == "Topic" else ("Response" if environment == "Financial Well-Being" else group_label)
            legend_dim_label = axis_legend if axis_legend == "Topic" else ("Response" if environment == "Financial Well-Being" else group_label)

            chart_data["x"] = chart_data[x_col]
            color_col = legend_col
            x_label = x_dim_label
            title = f"P-Fin 8: {_metric_label} — {single_group_value}" if single_group_value else f"P-Fin 8: {_metric_label} — {x_dim_label} × {legend_dim_label}"

        else:
            topics_map = {k: v for k, v in TOPIC_CAT3_NAMES.items() if k in selected_topics}
            chart_data = prepare_topic_cat3_data(df, topics_map, group_col, group_label)
            # Filter to selected response categories
            if selected_response_cats:
                chart_data = chart_data[chart_data["response_category"].isin(selected_response_cats)]
            hover_mode = "cat3"
            if selected_response_cats and len(selected_response_cats) == 1:
                y_label = f"% {selected_response_cats[0]}"
            else:
                y_label = "% of Respondents"

            # Assign axes for 3 dimensions
            x_col = dim_to_col(axis_x, "cat3")
            legend_col = dim_to_col(axis_legend, "cat3")
            facet_dim = axis_facet
            facet_col = dim_to_col(axis_facet, "cat3") if axis_facet else None

            # For stacked bar and pie, Response Category must be in the legend (color)
            # If it's on the x-axis, swap x and legend
            if chart_type in ["Stacked Bar Chart", "Horizontal Stacked Bar Chart", "Pie Chart"] and x_col == "response_category":
                x_col, legend_col = legend_col, x_col
                axis_x, axis_legend = axis_legend, axis_x

            x_dim_label = "Topic" if axis_x == "Topic" else ("Response Category" if axis_x == "Response Category" else ("Response" if environment == "Financial Well-Being" else group_label))
            legend_dim_label = "Topic" if axis_legend == "Topic" else ("Response Category" if axis_legend == "Response Category" else ("Response" if environment == "Financial Well-Being" else group_label))

            chart_data["x"] = chart_data[x_col]
            color_col = legend_col
            # Only facet if the facet dimension has multiple values
            if facet_col and facet_col in chart_data.columns and chart_data[facet_col].nunique() > 1:
                use_facet = facet_col
            else:
                use_facet = None
            x_label = x_dim_label
            title = f"P-Fin 8: Response Distribution — {single_group_value}" if single_group_value else f"P-Fin 8: Response Distribution by {group_label}"

            # Set category orders for response_category if used
            if "response_category" in [x_col, legend_col, facet_col]:
                _full_cat_order = ["Correct", "Incorrect", "Don't Know"]
                category_orders["response_category"] = [c for c in _full_cat_order if not selected_response_cats or c in selected_response_cats]

    else:
        selected_range = config["selected_range"]
        if not selected_range:
            return None, None, None, None, None
        chart_data = prepare_total_correct_data(df, group_col, score_range=selected_range)
        hover_mode = "total_correct"
        y_label = "% of Respondents"

        # Assign axes
        x_col = dim_to_col(axis_x)
        legend_col = dim_to_col(axis_legend)
        x_dim_label = "Number Correct" if axis_x == "Number Correct" else ("Response" if environment == "Financial Well-Being" else group_label)
        legend_dim_label = "Number Correct" if axis_legend == "Number Correct" else ("Response" if environment == "Financial Well-Being" else group_label)

        chart_data["x"] = chart_data[x_col]
        color_col = legend_col
        x_label = x_dim_label
        title = f"P-Fin 8: Distribution of Number Correct — {single_group_value}" if single_group_value else f"P-Fin 8: Distribution of Number Correct — {x_dim_label} × {legend_dim_label}"

        # Ensure score order
        if not chart_data.empty:
            score_labels = [TOTAL_CORRECT_LABELS[i] for i in sorted(selected_range) if i in TOTAL_CORRECT_LABELS]
            category_orders["score_label"] = score_labels
            if x_col == "score_label":
                category_orders["x"] = score_labels

    if chart_data is None or chart_data.empty:
        st.warning("No data available for the selected combination. Please adjust your filters.")
        return None, None, None, None, None

    # Set legend label
    legend_label_text = None
    if environment == "Financial Well-Being" and legend_col == "group_value":
        legend_label_text = "Response"
    elif legend_col == "response_category":
        legend_label_text = "Response Category"
    elif legend_col == "topic":
        legend_label_text = "Topic"
    elif legend_col == "score_label":
        legend_label_text = "Number Correct"
    else:
        legend_label_text = group_label

    # Create chart or table
    fig = None
    if chart_type != "Table":
        # Compute n_legend_groups from actual data
        actual_n_legend_groups = chart_data[color_col].nunique() if color_col in chart_data.columns else 1

        # Determine pie chart slice column (parts-of-whole dimension)
        pie_names = None
        if chart_type == "Pie Chart":
            if analysis_type == "Topic Bucket" and view_mode and "3-Category" in view_mode:
                pie_names = "response_category"
            elif analysis_type == "Number Correct":
                pie_names = "score_label"

        fig = create_chart(chart_data, chart_type, title, x_label, y_label, color_col,
                           category_orders, group_label=legend_label_text, hover_mode=hover_mode,
                           legend_label=legend_label_text, facet_col=use_facet,
                           n_legend_groups=actual_n_legend_groups, pie_names_col=pie_names,
                           show_pct_labels=config.get("show_pct_labels", False))

    # Generate note
    note = generate_note(
        environment=environment,
        analysis_type=analysis_type,
        view_mode=view_mode,
        selected_topics=config["selected_topics"],
        selected_range=config["selected_range"],
        analysis_variable=config.get("analysis_variable"),
        subgroups=config.get("subgroups"),
        n_obs=len(df),
        dataset_name=dataset_name,
        year_range=config.get("selected_years"),
    )

    # Run sanity checks
    checks = run_sanity_checks(df, "survey_weight", chart_data, environment, analysis_col)

    return fig, note, checks, chart_data, title


# ==============================================================================
# DEBUG PANEL
# ==============================================================================
def render_debug_panel(checks):
    if not DEBUG_MODE or checks is None:
        return

    with st.expander("🔧 Debug: Validation Panel", expanded=True):
        col1, col2, col3 = st.columns(3)

        with col1:
            st.markdown("**✅ Passed**")
            for item in checks["passed"]:
                st.write(f"• {item}")

        with col2:
            st.markdown("**⚠️ Warnings**")
            if checks["warnings"]:
                for item in checks["warnings"]:
                    st.warning(item)
            else:
                st.write("No warnings")

        with col3:
            st.markdown("**❌ Errors**")
            if checks["errors"]:
                for item in checks["errors"]:
                    st.error(item)
            else:
                st.write("No errors")


# ==============================================================================
# MAIN APP
# ==============================================================================
def main():
    st.set_page_config(
        page_title="P-Fin 8 Data Exploration Tool",
        page_icon=None,
        layout="wide",
    )

    # Load data
    try:
        df_years = load_all_years()
        df_genpop = load_genpop()
    except FileNotFoundError as e:
        st.error(
            f"Data file not found: {e}. Please ensure 'allYearsPFin8.xlsx' and "
            f"'PFin2026_GenPop.xlsx' are in the directory: {DATA_DIR.resolve()}"
        )
        return

    # Render sidebar and get config
    config = render_sidebar(df_years, df_genpop)

    # Main content
    st.title("P-Fin 8 Data Exploration Tool")
    st.markdown(
        "This tool was designed to help users explore the P-Fin 8 Index data through customizable visual "
        "analysis. All visual analysis presented in the tool is based on weighted data so that the results "
        "are representative of the national population. To create a visual, please work through the toolbar "
        "in the following order: **Analysis Type → Analysis Type Filters → Exploration Type → Variable "
        "Selection → Variable Filters → Chart Type → Axis Assignment**. If you plan to use any visual "
        "created with this tool in your own work, please cite The TIAA Institute–GFLEC Personal Finance "
        "Index. We hope you enjoy exploring the data."
    )

    # Custom CSS
    st.markdown("""
        <style>
        [data-testid="stButton"] button:hover {
            text-decoration: underline !important;
            color: #1f4e79 !important;
        }
        [data-testid="stButton"] button p {
            font-weight: 800 !important;
        }
        /* Hide anchor links on headers */
        h1 a, h2 a, h3 a, h4 a, h5 a, h6 a,
        [data-testid="stMarkdown"] h1 a,
        [data-testid="stMarkdown"] h2 a,
        [data-testid="stMarkdown"] h3 a {
            display: none !important;
        }
        .stMainBlockContainer h1 a,
        .stMainBlockContainer h2 a,
        .stMainBlockContainer h3 a {
            display: none !important;
        }
        </style>
    """, unsafe_allow_html=True)

    st.markdown('<hr style="margin-top:0.5rem;margin-bottom:0;">', unsafe_allow_html=True)

    # Run analysis
    fig, note, checks, chart_data, chart_title = run_analysis(config, df_years, df_genpop)

    # Display chart or table
    if config["chart_type"] == "Table" and chart_data is not None and not chart_data.empty:

        # Build a pivoted display table
        # Rows = what's on x-axis, Columns = what's in legend
        axis_x = config.get("axis_x", "")
        axis_legend = config.get("axis_legend", "")
        analysis_type = config.get("analysis_type", "")
        view_mode = config.get("view_mode", "")

        # Map axis selections to data column names
        def axis_to_col(axis_name):
            if axis_name == "Topic":
                return "topic"
            elif axis_name == "Number Correct":
                return "score_label"
            elif axis_name == "Response Category":
                return "response_category"
            else:
                return "group_value"

        row_col = axis_to_col(axis_x)
        col_col = axis_to_col(axis_legend)

        # For 3-category with facet, we need to handle the third dimension
        axis_facet = config.get("axis_facet")
        facet_col_name = axis_to_col(axis_facet) if axis_facet else None

        chart_data["percentage"] = chart_data["percentage"].round(2)

        # Get response counts per row group
        n_counts = chart_data.groupby(row_col)["n"].first()

        if facet_col_name:
            # 3 dimensions: pivot with multi-level columns (facet × legend)
            pivot_df = chart_data.pivot_table(
                index=row_col, columns=[facet_col_name, col_col],
                values="percentage", aggfunc="first"
            )
            # Flatten multi-level column names
            pivot_df.columns = [f"{f} — {c}" for f, c in pivot_df.columns]
            # Capture group structure (facet -> [categories]) for grouped display
            facet_groups = {}
            for col in pivot_df.columns:
                parts = col.split(" — ", 1)
                if len(parts) == 2:
                    f_name, c_name = parts
                    if f_name not in facet_groups:
                        facet_groups[f_name] = []
                    facet_groups[f_name].append(c_name)
        else:
            # 2 dimensions: simple pivot
            pivot_df = chart_data.pivot_table(
                index=row_col, columns=col_col,
                values="percentage", aggfunc="first"
            )
            facet_groups = None

        # Add response count column
        pivot_df["Response Count"] = n_counts

        # Reset index so row variable becomes a column
        pivot_df = pivot_df.reset_index()

        # Rename the index column for display
        row_label = axis_x if axis_x in ["Topic", "Number Correct", "Response Category"] else config.get("group_dim_label", "Group")
        pivot_df = pivot_df.rename(columns={row_col: row_label})

        # Maintain category order if applicable
        if row_col == "score_label":
            score_order = [TOTAL_CORRECT_LABELS[i] for i in range(9)]
            available = [s for s in score_order if s in pivot_df[row_label].values]
            pivot_df[row_label] = pd.Categorical(pivot_df[row_label], categories=available, ordered=True)
            pivot_df = pivot_df.sort_values(row_label)

        pivot_df = pivot_df.set_index(row_label)

        # Format percentage columns with 2 decimals and % sign, keep Response Count as integer
        pct_cols = [c for c in pivot_df.columns if c != "Response Count"]
        pivot_df["Response Count"] = pivot_df["Response Count"].astype(int)
        for col in pct_cols:
            pivot_df[col] = pivot_df[col].apply(lambda v: f"{v:.2f}%" if pd.notna(v) else "")

        # Title and export links
        import base64
        from io import BytesIO

        if facet_groups:
            # Two-row CSV: row 1 = group names (with blanks for extra cols), row 2 = category names
            import csv as _csv
            import io as _io
            _csv_buf = _io.StringIO()
            _writer = _csv.writer(_csv_buf)
            # Header row 1
            _r1 = [pivot_df.index.name or ""]
            for _facet, _cats in facet_groups.items():
                _r1.append(_facet)
                _r1.extend([""] * (len(_cats) - 1))
            _r1.append("Response Count")
            _writer.writerow(_r1)
            # Header row 2
            _r2 = [""]
            for _facet, _cats in facet_groups.items():
                _r2.extend(_cats)
            _r2.append("")
            _writer.writerow(_r2)
            # Data rows
            for _idx, _row in pivot_df.iterrows():
                _dr = [_idx]
                for _facet, _cats in facet_groups.items():
                    for _cat in _cats:
                        _dr.append(_row.get(f"{_facet} — {_cat}", ""))
                _dr.append(_row.get("Response Count", ""))
                _writer.writerow(_dr)
            csv_data = _csv_buf.getvalue()
        else:
            csv_data = pivot_df.to_csv(index=True)
        csv_b64 = base64.b64encode(csv_data.encode()).decode()

        if facet_groups:
            # Excel with merged group headers via openpyxl
            from openpyxl import Workbook as _Workbook
            from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Border, Side as _Side
            from openpyxl.utils import get_column_letter as _gcl
            _wb = _Workbook()
            _ws = _wb.active
            _hdr_fill = _Fill("solid", fgColor="1F4E79")
            _hdr_font = _Font(color="FFFFFF", bold=True)
            _sub_fill = _Fill("solid", fgColor="D0E4F7")
            _sub_font = _Font(bold=True)
            _ctr = _Align(horizontal="center", vertical="center", wrap_text=True)
            _grp_side = _Side(style="thin", color="AAAAAA")
            _no_side = _Side(style=None)

            def _grp_border(is_first, is_last):
                """Light gray left/right border on outermost cells of a group."""
                return _Border(
                    left=_grp_side if is_first else _no_side,
                    right=_grp_side if is_last else _no_side,
                )

            # Build a mapping: column index -> (is_first_in_group, is_last_in_group)
            # so we can apply borders uniformly across header and data rows.
            _grp_border_map = {}  # col_index (1-based) -> border object
            _ci = 1
            # Index col: right border acts as left boundary of first group
            _grp_border_map[_ci] = _Border(right=_grp_side)
            _ci += 1
            _facet_list = list(facet_groups.items())
            for _fi, (_facet, _cats) in enumerate(_facet_list):
                for _ki, _cat in enumerate(_cats):
                    _is_first = (_ki == 0)
                    _is_last = (_ki == len(_cats) - 1)
                    _grp_border_map[_ci] = _grp_border(_is_first, _is_last)
                    _ci += 1
            # Response Count col: left border acts as right boundary of last group
            _grp_border_map[_ci] = _Border(left=_grp_side)

            # Index column spanning rows 1–2
            _ci = 1
            _c = _ws.cell(row=1, column=_ci, value=pivot_df.index.name or "")
            _c.fill, _c.font, _c.alignment = _hdr_fill, _hdr_font, _ctr
            _c.border = _grp_border_map.get(_ci, _Border())
            _ws.merge_cells(start_row=1, start_column=_ci, end_row=2, end_column=_ci)
            _ci += 1
            # Group headers (merged across categories) + category sub-headers
            for _facet, _cats in facet_groups.items():
                _start = _ci
                _c = _ws.cell(row=1, column=_ci, value=_facet)
                _c.fill, _c.font, _c.alignment = _hdr_fill, _hdr_font, _ctr
                _c.border = _grp_border_map.get(_ci, _Border())
                if len(_cats) > 1:
                    _ws.merge_cells(start_row=1, start_column=_start, end_row=1, end_column=_start + len(_cats) - 1)
                for _ki2, _cat in enumerate(_cats):
                    _c2 = _ws.cell(row=2, column=_ci, value=_cat)
                    _c2.fill, _c2.font, _c2.alignment = _sub_fill, _sub_font, _ctr
                    # Every sub-header cell gets thin gray dividers on both sides;
                    # outer group edges use the group-boundary border from the map.
                    _grp_b = _grp_border_map.get(_ci, _Border())
                    _c2.border = _Border(
                        left=_grp_b.left if _grp_b.left and _grp_b.left.style else _grp_side,
                        right=_grp_b.right if _grp_b.right and _grp_b.right.style else _grp_side,
                        top=_grp_side,
                        bottom=_grp_side,
                    )
                    _ci += 1
            # Response Count spanning rows 1–2
            _c = _ws.cell(row=1, column=_ci, value="Response Count")
            _c.fill, _c.font, _c.alignment = _hdr_fill, _hdr_font, _ctr
            _c.border = _grp_border_map.get(_ci, _Border())
            _ws.merge_cells(start_row=1, start_column=_ci, end_row=2, end_column=_ci)
            def _col_border(col_idx):
                """Gray divider on both sides of every data cell; group edges keep their border."""
                _grp_b = _grp_border_map.get(col_idx, _Border())
                return _Border(
                    left=_grp_b.left if _grp_b.left and _grp_b.left.style else _grp_side,
                    right=_grp_b.right if _grp_b.right and _grp_b.right.style else _grp_side,
                )

            # Data rows
            _n_data_rows = len(pivot_df)
            for _rn, (_idx, _row) in enumerate(pivot_df.iterrows(), start=3):
                _bg = "F9F9F9" if (_rn - 3) % 2 == 0 else "FFFFFF"
                _rf = _Fill("solid", fgColor=_bg)
                _ci2 = 1
                _c = _ws.cell(row=_rn, column=_ci2, value=_idx)
                _c.font, _c.fill = _Font(bold=True), _rf
                _c.border = _col_border(_ci2)
                _ci2 += 1
                for _facet, _cats in facet_groups.items():
                    for _cat in _cats:
                        _c = _ws.cell(row=_rn, column=_ci2, value=_row.get(f"{_facet} — {_cat}", ""))
                        _c.alignment, _c.fill = _Align(horizontal="center"), _rf
                        _c.border = _col_border(_ci2)
                        _ci2 += 1
                _c = _ws.cell(row=_rn, column=_ci2, value=_row.get("Response Count", ""))
                _c.alignment, _c.fill = _Align(horizontal="center"), _rf
                _c.border = _col_border(_ci2)
            # Auto-width columns
            for _col in _ws.columns:
                _ml = max((len(str(_cell.value)) for _cell in _col if _cell.value), default=6)
                _ws.column_dimensions[_gcl(_col[0].column)].width = min(30, _ml + 3)
            _ws.row_dimensions[1].height = 30
            _ws.row_dimensions[2].height = 20
            excel_buffer = BytesIO()
            _wb.save(excel_buffer)
            excel_buffer.seek(0)
        else:
            excel_buffer = BytesIO()
            pivot_df.to_excel(excel_buffer, index=True, engine="openpyxl")
            excel_buffer.seek(0)
        xlsx_b64 = base64.b64encode(excel_buffer.getvalue()).decode()

        # Build HTML table string for facet case (used for display and html2canvas PNG)
        if facet_groups:
            row_lbl = pivot_df.index.name or ""
            _th1 = (
                f'<th rowspan="2" style="background:#1f4e79;color:white;'
                f'text-align:center !important;padding:8px 12px;border:1px solid #ccc;">{row_lbl}</th>'
            )
            for facet, cats in facet_groups.items():
                _th1 += (
                    f'<th colspan="{len(cats)}" style="background:#1f4e79;color:white;'
                    f'text-align:center !important;padding:8px 12px;border:1px solid #ccc;">{facet}</th>'
                )
            _th1 += (
                '<th rowspan="2" style="background:#1f4e79;color:white;'
                'text-align:center !important;padding:8px 12px;border:1px solid #ccc;">Response Count</th>'
            )
            _th2 = ""
            for facet, cats in facet_groups.items():
                for cat in cats:
                    _th2 += (
                        f'<th style="background:#d0e4f7;color:black;'
                        f'text-align:center !important;padding:6px 10px;border:1px solid #ccc;">{cat}</th>'
                    )
            _tbody = ""
            for i, (idx, row) in enumerate(pivot_df.iterrows()):
                bg = "#f9f9f9" if i % 2 == 0 else "white"
                _tbody += f'<tr style="background:{bg}">'
                _tbody += f'<td style="padding:6px 10px;border:1px solid #eee;font-weight:bold;">{idx}</td>'
                for facet, cats in facet_groups.items():
                    for cat in cats:
                        val = row.get(f"{facet} — {cat}", "")
                        _tbody += f'<td style="text-align:center;padding:6px 10px;border:1px solid #eee;">{val}</td>'
                _tbody += (
                    f'<td style="text-align:center;padding:6px 10px;border:1px solid #eee;">'
                    f'{row.get("Response Count", "")}</td>'
                )
                _tbody += "</tr>"
            table_html_inner = (
                '<style>.pfin8-table thead th { text-align: center !important; }</style>'
                '<table class="pfin8-table" style="border-collapse:collapse;width:100%;font-family:sans-serif;font-size:14px;">'
                f'<thead><tr>{_th1}</tr><tr>{_th2}</tr></thead>'
                f'<tbody>{_tbody}</tbody>'
                '</table>'
            )
        else:
            table_html_inner = None

        # Generate table as PNG using Plotly go.Table (non-facet only;
        # facet tables use client-side html2canvas via st.components)
        table_png_available = False
        table_png_b64 = ""
        try:
            if not facet_groups:
                # Non-facet: single header row via go.Table
                header_vals = [pivot_df.index.name or ""] + list(pivot_df.columns)
                cell_vals = [[str(v) for v in pivot_df.index]] + [
                    [str(v) for v in pivot_df[col]] for col in pivot_df.columns
                ]
                table_fig = go.Figure(data=[go.Table(
                    header=dict(
                        values=[f"<b>{h}</b>" for h in header_vals],
                        fill_color="#636EFA",
                        font=dict(color="white", size=13),
                        align="center",
                    ),
                    cells=dict(
                        values=cell_vals,
                        fill_color=[["#f9f9f9", "white"] * (len(pivot_df) // 2 + 1)],
                        font=dict(size=12),
                        align=["left"] + ["center"] * len(pivot_df.columns),
                    ),
                )])
                n_rows = len(pivot_df)
                n_cols = len(header_vals)
                max_header_len = max(len(str(h)) for h in header_vals)
                col_width = max(150, max_header_len * 12)
                table_fig.update_layout(
                    title=f"{chart_title}<br><sup><span style='color:gray;font-size:11px'>Source: TIAA G-Flec Personal Finance Index</span></sup>",
                    title_font=dict(size=16),
                    width=max(900, n_cols * col_width),
                    height=max(400, 80 + n_rows * 35),
                    margin=dict(l=10, r=10, t=50, b=10),
                )
            table_png_bytes = table_fig.to_image(format="png", scale=2)
            table_png_b64 = base64.b64encode(table_png_bytes).decode()
            table_png_available = True
        except Exception:
            pass

        title_col, dl_col = st.columns([7, 3])
        with title_col:
            st.markdown(f"### {chart_title}")
            st.caption("Source: TIAA G-Flec Personal Finance Index")
        with dl_col:
            if facet_groups and table_html_inner:
                # Facet tables: client-side PNG via html2canvas + CSV + Excel
                import streamlit.components.v1 as components
                component_html = (
                    '<!DOCTYPE html><html><head>'
                    '<script src="https://html2canvas.hertzen.com/dist/html2canvas.min.js"></script>'
                    '<style>'
                    '* { box-sizing: border-box; }'
                    'body { margin: 0; padding: 0; font-family: sans-serif; overflow: hidden; }'
                    '#capture { position: absolute; left: -9999px; top: 0;'
                    ' background: white; padding: 20px; width: max-content; }'
                    '#capture h3 { color: black; margin: 0 0 12px 0; font-size: 18px; }'
                    '.dl-bar { text-align: right; padding: 8px 0; font-size: 0.875rem; white-space: nowrap; }'
                    'a { color: #1f77b4; text-decoration: underline; cursor: pointer; }'
                    '</style></head><body>'
                    f'<div id="capture"><h3>{chart_title}</h3>{table_html_inner}</div>'
                    '<div class="dl-bar">Download: '
                    '<a id="png-btn" href="#">PNG</a> | '
                    f'<a href="data:text/csv;base64,{csv_b64}" download="pfin8_table.csv">CSV</a> | '
                    f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{xlsx_b64}" download="pfin8_table.xlsx">Excel</a>'
                    '</div>'
                    '<script>'
                    'document.getElementById("png-btn").addEventListener("click", function(e) {'
                    '  e.preventDefault();'
                    '  var el = document.getElementById("capture");'
                    '  html2canvas(el, { scale: 2, backgroundColor: "#ffffff", logging: false, useCORS: true })'
                    '  .then(function(canvas) {'
                    '    var a = document.createElement("a");'
                    '    a.href = canvas.toDataURL("image/png");'
                    '    a.download = "pfin8_table.png";'
                    '    document.body.appendChild(a); a.click(); document.body.removeChild(a);'
                    '  });'
                    '});'
                    '</script></body></html>'
                )
                components.html(component_html, height=40, scrolling=False)
            else:
                if table_png_available:
                    download_html = (
                        f'<div style="text-align:right; font-size:0.9rem; padding:8px 0;">'
                        f'Download: '
                        f'<a href="data:image/png;base64,{table_png_b64}" download="pfin8_table.png" '
                        f'style="color:#1f77b4; text-decoration:underline;">PNG</a>'
                        f' | '
                        f'<a href="data:text/csv;base64,{csv_b64}" download="pfin8_table.csv" '
                        f'style="color:#1f77b4; text-decoration:underline;">CSV</a>'
                        f' | '
                        f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{xlsx_b64}" download="pfin8_table.xlsx" '
                        f'style="color:#1f77b4; text-decoration:underline;">Excel</a>'
                        f'</div>'
                    )
                else:
                    download_html = (
                        f'<div style="text-align:right; font-size:0.9rem; padding:8px 0;">'
                        f'Download: '
                        f'<a href="data:text/csv;base64,{csv_b64}" download="pfin8_table.csv" '
                        f'style="color:#1f77b4; text-decoration:underline;">CSV</a>'
                        f' | '
                        f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{xlsx_b64}" download="pfin8_table.xlsx" '
                        f'style="color:#1f77b4; text-decoration:underline;">Excel</a>'
                        f'</div>'
                    )
                st.markdown(download_html, unsafe_allow_html=True)

        if facet_groups and table_html_inner:
            st.markdown(
                f'<div style="overflow-x:auto;font-size:0.85rem;">{table_html_inner}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.table(pivot_df)

        # Display sample size warnings inline
        if checks and checks["warnings"]:
            for warning in checks["warnings"]:
                st.caption(warning)

        # Display note
        if note:
            st.markdown("---")
            import re
            note_html = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', note)
            st.markdown(f'<div style="color: black; font-size: 0.85rem;">{note_html}</div>', unsafe_allow_html=True)

        # Debug panel
        render_debug_panel(checks)

    elif fig:
        # Try to generate PNG
        png_available = False
        png_bytes = None
        try:
            png_bytes = fig.to_image(format="png", width=2000, height=fig.layout.height or 600, scale=2)
            png_available = True
        except Exception:
            pass

        # Build HTML download data
        import base64
        html_data = fig.to_html(include_plotlyjs="cdn")
        html_b64 = base64.b64encode(html_data.encode()).decode()

        if png_available:
            png_b64 = base64.b64encode(png_bytes).decode()
            download_html = (
                f'<div style="text-align:right; font-size:0.9rem; padding:8px 0;">'
                f'Download: '
                f'<a href="data:image/png;base64,{png_b64}" download="pfin8_chart.png" '
                f'style="color:#1f77b4; text-decoration:underline;">PNG</a>'
                f' | '
                f'<a href="data:text/html;base64,{html_b64}" download="pfin8_chart.html" '
                f'style="color:#1f77b4; text-decoration:underline;">HTML</a>'
                f'</div>'
            )
        else:
            download_html = (
                f'<div style="text-align:right; font-size:0.9rem; padding:8px 0;">'
                f'Download: '
                f'<a href="data:text/html;base64,{html_b64}" download="pfin8_chart.html" '
                f'style="color:#1f77b4; text-decoration:underline;">HTML</a>'
                f'</div>'
            )
        st.markdown(download_html, unsafe_allow_html=True)

        st.plotly_chart(fig, use_container_width=True)

        # Display sample size warnings inline
        if checks and checks["warnings"]:
            for warning in checks["warnings"]:
                st.caption(warning)

        # Display note
        if note:
            st.markdown("---")
            import re
            note_html = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', note)
            st.markdown(f'<div style="color: black; font-size: 0.85rem;">{note_html}</div>', unsafe_allow_html=True)

        # Debug panel
        render_debug_panel(checks)


if __name__ == "__main__":
    main()
